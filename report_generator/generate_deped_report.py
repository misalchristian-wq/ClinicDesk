#!/usr/bin/env python3
"""
generate_deped_report.py
Fills the official DepEd Part IX xlsx template with ClinicDesk data for a
school year, WITHOUT altering the template's design or formulas. Only blank
leaf cells are written; DepEd SUM() totals recalculate on their own.

Usage: python3 generate_deped_report.py <template> <output> <school_year> <data.json>
data.json: {"students":[...], "immunization":[...], "deworming":[...]}
Sheets filled: Table 1.B (nutrition), Table 1.A (immunization), Table 1.C (deworming),
              Table 1.D (WIFA), Box 1, Boxes 5 & 6.
"""
import sys, json
import datetime
import openpyxl

def norm_grade(g):
    if g is None: return None
    raw = str(g).strip().lower()
    if "kinder" in raw or raw == "k": return "Kinder"
    if "sned" in raw or "non-graded" in raw or "nongraded" in raw: return "SNEd"
    s = raw.replace("grade", "").strip()
    return "Grade " + s if s.isdigit() else None

def norm_sex(x):
    return "M" if str(x).strip().lower().startswith("m") else "F"

def truthy(v):
    return str(v).strip().lower() in ("1", "yes", "y", "true", "t")

def put(ws, col, row, val):
    if val:
        ws[f"{col}{row}"] = val

# ---- Table 1.B Nutrition ----
NUTRI_SHEET = "IX. Table 1.A & 1.B"
ELEM_ROWS = {"Normal":22,"Obese":23,"Overweight":24,"Severely Wasted":25,"Wasted":26}
ELEM_COLS = {"Kinder":("I","L"),"Grade 1":("O","R"),"Grade 2":("U","X"),"Grade 3":("AA","AD"),
             "Grade 4":("AG","AJ"),"Grade 5":("AM","AP"),"Grade 6":("AS","AV"),"SNEd":("AY","BB")}
SEC_ROWS = {"Normal":33,"Obese":34,"Overweight":35,"Severely Wasted":36,"Wasted":37}
JHS_COLS = {"Grade 7":("I","L"),"Grade 8":("O","R"),"Grade 9":("U","X"),"Grade 10":("AA","AD")}
SHS_COLS = {"Grade 11":("AM","AP"),"Grade 12":("AS","AV")}

def norm_bmi(c):
    s = str(c or "").strip().lower()
    if "severely" in s and "wast" in s: return "Severely Wasted"
    if "wast" in s: return "Wasted"
    if "obese" in s: return "Obese"
    if "overweight" in s: return "Overweight"
    if "normal" in s: return "Normal"
    return None

def fill_nutrition(ws, students, sy):
    counts = {}
    for r in students:
        if str(r.get("school_year","")).strip() != sy: continue
        grade = norm_grade(r.get("grade_level")); cat = norm_bmi(r.get("bmi_category"))
        if not grade or not cat: continue
        sex = norm_sex(r.get("sex"))
        counts[(grade,sex,cat)] = counts.get((grade,sex,cat),0)+1
    total = 0
    for (grade,sex,cat),n in counts.items():
        if grade in ELEM_COLS: rows,cols = ELEM_ROWS,ELEM_COLS
        elif grade in JHS_COLS: rows,cols = SEC_ROWS,JHS_COLS
        elif grade in SHS_COLS: rows,cols = SEC_ROWS,SHS_COLS
        else: continue
        if cat not in rows: continue
        mcol,fcol = cols[grade]
        put(ws, mcol if sex=="M" else fcol, rows[cat], n); total += n
    return total

# ---- Table 1.A Immunization (TOTAL learner columns) ----
IMMUN_ROWS = {"Measles Rubella":11,"Tetanus Diphtheria":12,"Human Papilloma Virus":13}
IMMUN_COLS = {"Grade 1":("L","Q"),"Grade 4":("AD","AD"),"Grade 7":("AQ","AV")}

def norm_vaccine(v):
    s = str(v or "").strip().lower()
    if "measles" in s or s == "mr" or "rubella" in s: return "Measles Rubella"
    if "tetanus" in s or s == "td" or "diphther" in s: return "Tetanus Diphtheria"
    if "hpv" in s or "papilloma" in s: return "Human Papilloma Virus"
    return None

def fill_immunization(ws, immun, sy):
    counts = {}
    for r in immun:
        if not truthy(r.get("immunized",1)): continue
        vac = norm_vaccine(r.get("vaccine")); grade = norm_grade(r.get("grade_level"))
        if not vac or not grade: continue
        sex = norm_sex(r.get("sex"))
        counts[(vac,grade,sex)] = counts.get((vac,grade,sex),0)+1
    total = 0
    for (vac,grade,sex),n in counts.items():
        row = IMMUN_ROWS.get(vac); cols = IMMUN_COLS.get(grade)
        if not row or not cols: continue
        mcol,fcol = cols
        put(ws, mcol if sex=="M" else fcol, row, n); total += n
    return total

# ---- Table 1.C Deworming ----
DEWORM_SHEET = "IX. Table 1.C & 1.D"
DEWORM_ROWS = {"Kinder":11,"Grade 1":12,"Grade 2":13,"Grade 3":14,"Grade 4":15,"Grade 5":16,
               "Grade 6":17,"SNEd":18,"Grade 7":21,"Grade 8":22,"Grade 9":23,"Grade 10":24,
               "Grade 11":27,"Grade 12":28}

def fill_deworming(ws, deworm, sy):
    sbfp = {}; other = {}
    for r in deworm:
        grade = norm_grade(r.get("grade_level"))
        if not grade: continue
        sex = norm_sex(r.get("sex"))
        if truthy(r.get("dewormed_sbfp")): sbfp[(grade,sex)] = sbfp.get((grade,sex),0)+1
        if truthy(r.get("dewormed_other")): other[(grade,sex)] = other.get((grade,sex),0)+1
    total = 0
    for (grade,sex),n in sbfp.items():
        row = DEWORM_ROWS.get(grade)
        if row: put(ws, "L" if sex=="M" else "Q", row, n); total += n
    for (grade,sex),n in other.items():
        row = DEWORM_ROWS.get(grade)
        if row: put(ws, "AA" if sex=="M" else "AF", row, n); total += n
    return total

# ---- Table 1.D WIFA (new) ----
def fill_wifa(ws, deworm, sy):
    # Group WIFA counts by grade and period (Jul-Sep, Jan-Mar) for females only
    wifa_counts = {}  # (grade, period) -> count
    for r in deworm:
        # Skip if not WIFA or not female
        if not truthy(r.get("wifa")):
            continue
        sex = norm_sex(r.get("sex"))
        if sex != "F":
            continue
        grade = norm_grade(r.get("grade_level"))
        if not grade:
            continue
        # Determine period from wifa_date
        date_str = str(r.get("wifa_date", "")).strip()
        if not date_str:
            continue
        try:
            dt = datetime.datetime.strptime(date_str, "%Y-%m-%d")
            month = dt.month
            if 7 <= month <= 9:
                period = "jul_sep"
            elif 1 <= month <= 3:
                period = "jan_mar"
            else:
                continue
        except:
            continue
        key = (grade, period)
        wifa_counts[key] = wifa_counts.get(key, 0) + 1

    # Map grade to column letters (female only, single column)
    grade_cols = {
        "Grade 7": "Q",
        "Grade 8": "U",
        "Grade 9": "Y",
        "Grade 10": "AC",
        "Grade 11": "AP",
        "Grade 12": "AT",
    }

    # Row numbers per level and period
    rows = {
        "jhs": {"jul_sep": 35, "jan_mar": 37},
        "shs": {"jul_sep": 36, "jan_mar": 38},
    }

    total = 0
    for (grade, period), count in wifa_counts.items():
        col = grade_cols.get(grade)
        if not col: continue
        grade_num = int(grade.split()[-1]) if grade.startswith("Grade") else None
        if grade_num is None: continue
        level = "jhs" if grade_num <= 10 else "shs"
        row = rows[level].get(period)
        if row:
            put(ws, col, row, count)
            total += count
    return total

# ---- Box 1 (OKD & LHAS) ----
LHAS_SHEET = "IX. Box 1"
# Component -> row within each section (top-left of the merged data cell).
LHAS_COMP_ROW = {
    "ELEM": {"nutritional assessment":21,"health history":22,"vision screening":24,
             "hearing screening":25,"oral health":26,"cars":27,"rapid heeadsss":29},
    "JHS":  {"nutritional assessment":35,"health history":36,"vision screening":38,
             "hearing screening":39,"oral health":40,"cars":41,"rapid heeadsss":43},
    "SHS":  {"nutritional assessment":49,"health history":50,"vision screening":52,
             "hearing screening":53,"oral health":54,"cars":55,"rapid heeadsss":57},
}
# Column letters for each metric.
LHAS_COLS = {"masterlisted":"P","screened":"T","findings":"X",
             "referred_school":"AB","referred_lgu":"AF","referred_private":"AJ","referred_others":"AN"}

def norm_component(c):
    s = str(c or "").strip().lower()
    if "nutrition" in s: return "nutritional assessment"
    if "health history" in s or "hhi" in s: return "health history"
    if "vision" in s: return "vision screening"
    if "hearing" in s: return "hearing screening"
    if "oral" in s: return "oral health"
    if "cars" in s or "adolescent risk" in s: return "cars"
    if "heeadsss" in s or "heeadss" in s: return "rapid heeadsss"
    return None

def level_from_grade(grade):
    g = norm_grade(grade)
    if g in ("Kinder","Grade 1","Grade 2","Grade 3","Grade 4","Grade 5","Grade 6","SNEd"):
        return "ELEM"
    if g in ("Grade 7","Grade 8","Grade 9","Grade 10"):
        return "JHS"
    if g in ("Grade 11","Grade 12"):
        return "SHS"
    return None

def fill_lhas(ws, lhas, sy):
    # Aggregate per (level, component): sum each metric.
    agg = {}
    for r in lhas:
        comp = norm_component(r.get("screening_type"))
        level = level_from_grade(r.get("grade_level"))
        if not comp or not level:
            continue
        key = (level, comp)
        if key not in agg:
            agg[key] = {m: 0 for m in LHAS_COLS}
        agg[key]["masterlisted"]     += int(r.get("masterlisted", 0) or 0)
        agg[key]["screened"]         += int(r.get("screened", 0) or 0)
        agg[key]["findings"]         += int(r.get("findings", 0) or 0)
        agg[key]["referred_school"]  += int(r.get("referred_school", 0) or 0)
        agg[key]["referred_lgu"]     += int(r.get("referred_lgu", 0) or 0)
        agg[key]["referred_private"] += int(r.get("referred_private", 0) or 0)
        agg[key]["referred_others"]  += int(r.get("referred_others", 0) or 0)

    total = 0
    for (level, comp), metrics in agg.items():
        row = LHAS_COMP_ROW.get(level, {}).get(comp)
        if not row:
            continue
        for metric, col in LHAS_COLS.items():
            put(ws, col, row, metrics[metric])
            total += metrics[metric]
    return total

# ---- Boxes 5 & 6 (ARH + Tobacco) ----
BOX56_SHEET = "IX. Boxes 5 & 6"
# Pregnant learners: grade -> column (single count column per grade, merged).
ARH_GRADE_COLS = {
    "Grade 4": "P", "Grade 5": "S", "Grade 6": "V",
    "Grade 7": "AB", "Grade 8": "AE", "Grade 9": "AH", "Grade 10": "AK",
    "Grade 11": "AQ", "Grade 12": "AT",
}
ARH_ROWS = {"in school": 10, "adm": 11}  # In School = row 10, ADM = row 11

def fill_box5_box6(ws, arh_list, peer_educators, tobacco_list):
    filled = 0
    # --- Box 5: pregnant learners by grade + delivery mode ---
    counts = {}  # (grade, mode_row) -> total
    for r in arh_list:
        grade = norm_grade(r.get("grade_level"))
        if grade not in ARH_GRADE_COLS:
            continue
        mode = str(r.get("delivery_mode", "")).strip().lower()
        if "adm" in mode or "alternative" in mode:
            row = ARH_ROWS["adm"]
        else:
            row = ARH_ROWS["in school"]  # blank/in-school both go here
        total = int(r.get("total", 1) or 1)
        counts[(grade, row)] = counts.get((grade, row), 0) + total

    for (grade, row), n in counts.items():
        put(ws, ARH_GRADE_COLS[grade], row, n)
        filled += n

    # Peer educators -> AK15
    if peer_educators:
        put(ws, "AK", 15, int(peer_educators))

    # --- Box 6: tobacco brought / referred by level ---
    # Brought tobacco table: row 39 (brought), row 41 (referred).
    # Columns: Elementary=S, JHS=Z, SHS=AG.
    lvl_col = {"elem": "S", "jhs": "Z", "shs": "AG"}
    for t in tobacco_list:
        grp = t.get("level_group", "jhs")
        col = lvl_col.get(grp)
        if not col:
            continue
        put(ws, col, 39, int(t.get("brought", 0)))
        put(ws, col, 41, int(t.get("referred", 0)))
        filled += int(t.get("brought", 0))
    return filled

def main():
    if len(sys.argv) < 5:
        print(json.dumps({"success":False,"message":"usage: template output school_year data.json"})); sys.exit(1)
    template, output, sy, data_path = sys.argv[1:5]
    with open(data_path, encoding="utf-8") as f:
        data = json.load(f)
    students = data.get("students", []); immun = data.get("immunization", []); deworm = data.get("deworming", [])
    arh = data.get("arh", []); peer_educators = data.get("peer_educators", 0); tobacco = data.get("tobacco", [])
    lhas = data.get("lhas", [])
    wb = openpyxl.load_workbook(template)
    filled = {"nutrition":0,"immunization":0,"deworming":0,"box5_6":0,"lhas":0, "wifa":0}
    if NUTRI_SHEET in wb.sheetnames:
        filled["nutrition"] = fill_nutrition(wb[NUTRI_SHEET], students, sy)
        filled["immunization"] = fill_immunization(wb[NUTRI_SHEET], immun, sy)
    if DEWORM_SHEET in wb.sheetnames:
        filled["deworming"] = fill_deworming(wb[DEWORM_SHEET], deworm, sy)
        filled["wifa"] = fill_wifa(wb[DEWORM_SHEET], deworm, sy)   # <-- NEW
    if BOX56_SHEET in wb.sheetnames:
        filled["box5_6"] = fill_box5_box6(wb[BOX56_SHEET], arh, peer_educators, tobacco)
    if LHAS_SHEET in wb.sheetnames:
        filled["lhas"] = fill_lhas(wb[LHAS_SHEET], lhas, sy)
    wb.save(output)
    print(json.dumps({"success":True,"message":f"Report generated for {sy}.","filled":filled,"output":output}))

if __name__ == "__main__":
    main()